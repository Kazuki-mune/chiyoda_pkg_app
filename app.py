import pandas as pd
import openpyxl
import streamlit as st
import numpy as np

from openpyxl import load_workbook

# エクセルファイルをアップロード
uploaded_file = st.file_uploader("エクセルファイルをアップロードしてください", type=['xlsx', 'xls'])

if uploaded_file is not None:
    # アップロードされたファイルを読み込む
    df = pd.read_excel(uploaded_file)
    
    # ファイル名を保存（後で使用するため）
    original_filename = uploaded_file.name
    
    # 一時ファイルとして保存
    temp_path = f"temp_{original_filename}"
    with open(temp_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    st.write("✅ ファイルが正常に読み込まれました")
else:
    st.warning("⚠️ エクセルファイルをアップロードしてください")
    st.stop()

# すべての処理はsession_stateに格納して状態を維持する
if "saved" not in st.session_state:
    st.session_state.saved = False  # 初期状態では保存されていない


# 基本となる列の定義
base_columns = [
    '工場区分名', '機械工程分類区分名', '機械名', '製造日', '受注区分名', '得意先略称',
    '品種コード', '原反コード', '原反名称', '受注№', '品名', '版区分', '工程コード', '工程名',
    '作業者', '操業区分', '完了区分', '投入数', '製造数', 'ロス', '所要時間準備', '所要時間調色',
    '所要時間計画', '所要時間突発', '所要時間立合', '所要時間工程', '所要時間後始末',
    'ロス時間合計', '所要時間基礎', '所要時間合計', 'unnamed','色数', '調色回数', '手配数', '責任完了数',
    'リード紙Ｍ数', '巻M数1', '巻数1', '巻M数2', '巻数2', '通し（総）ｍ数', '生産金額+荷作り代',
    '印刷速度', 'ラミ工程', 'ラミ速度', '品種分類コード', '業種コード', '印刷ロス率(受注)',
    '印刷ロス率(実績)', '生産金額', '荷作り代', 'ロスM数', '原反巾', '原反受入単価', '工程単価'
]

# 分類列の定義
category_columns = ['56_製品分類No', '57_製品分類名']

# ロス列の作成
if 'ロス' not in df.columns:
    if '投入数' in df.columns and '製造数' in df.columns:
        df['ロス'] = df['投入数'] - df['製造数']
        st.write("🔧 『ロス』列を『投入数 - 製造数』で新たに作成しました。")
    else:
        st.warning("⚠️ 『ロス』列を作成できません（『投入数』『製造数』の列が不足）")

# unnamed列の作成
if 'unnamed' not in df.columns:
    df['unnamed'] = 0

# 分類列の追加
for col in category_columns:
    if col not in df.columns:
        df[col] = None

# 欠損列の確認
missing = [col for col in base_columns if col not in df.columns]
if missing:
    st.warning(f"⚠️ 以下の列が存在しません（空白のままになります）: {missing}")

# 列名の番号付け
present_base_columns = [col for col in base_columns if col in df.columns]
numbered_base_columns = [f"{i+1}_{col}" for i, col in enumerate(present_base_columns)]
rename_map = dict(zip(present_base_columns, numbered_base_columns))
df = df.rename(columns=rename_map)

# 所要時間基礎の計算
if '30_所要時間合計' in df.columns and '26_所要時間工程' in df.columns:
    df['29_所要時間基礎'] = df['30_所要時間合計'] - df['26_所要時間工程']
    st.write("🔧 『所要時間基礎』列を『所要時間合計 - 所要時間工程』で更新しました。")
else:
    st.warning("⚠️ 『所要時間基礎』列を更新できません（必要な列が不足）")

# 列の並び替え
final_columns = category_columns + numbered_base_columns
other_columns = [col for col in df.columns if col not in final_columns]
final_columns += other_columns
df = df[final_columns]

# 4_製造日を日付データに変換
if '4_製造日' in df.columns:
    df['4_製造日'] = pd.to_datetime(df['4_製造日'])
    st.write("🔧 『製造日』列を日付データに変換しました。")
# 品種コード230で麦茶５４を含み、試作を含まない行の製品分類Noを1に設定(麦茶54ｐ)
# 品種コード230で麦茶５４を含み、試作を含まない行の製品分類Noを1に設定(麦茶54ｐ)
mask = (df['7_品種コード'] == 230) & \
       (df['11_品名'].str.contains('麦茶５４', na=False)) 
df.loc[mask, '56_製品分類No'] = 1

#品種コード237で品名に以下の文字列を含む製品を製品分類Noを2に設定する(ＺＩＰ)
mask = (df['7_品種コード'] == 237) & \
        (df['11_品名'].str.contains('越後|セブン|香薫|あらびき|粗挽き|シャウエッセン'))
df.loc[mask, '56_製品分類No'] = 2

#品種コード239の製品を，製品分類No3に設定する(食品トレー)
mask = (df['7_品種コード'] == 239) 
df.loc[mask, '56_製品分類No'] = 3

#品種コード230で，得意先略称が伊藤園で，麦茶５４を含まない行の製品分類Noを4に設定(伊藤園製袋)
mask = (df['7_品種コード'] == 230) & \
       (df['6_得意先略称']=='伊藤園')&\
       (~df['11_品名'].str.contains('麦茶５４', na=False)) 
df.loc[mask, '56_製品分類No'] = 4

#品種コード236で得意先が伊藤園の製品を，製品分類Ｎｏ5に設定する．(伊藤園ロール)
mask = (df['7_品種コード'] == 236) & \
       (df['6_得意先略称']=='伊藤園')
df.loc[mask, '56_製品分類No'] = 5


#品種コード232で得意先がcus_patの5社であり，品名がpro_patである製品を製品分類Ｎｏ6に設定する．（ＩＭＬ)

customer_pattern = ['KISCO大阪本社', 'シーレックス', 'ﾘｽﾊﾟｯｸ㈱', '㈱ｱﾌﾟﾘｽ', '東洋科学']

product_pattern = [
    "【1枚抜き】メロンミックススムージ－21春",
    "【1枚抜き】梨ミックススムージー50g増量",
    "【1枚抜き専用】東洋科学様 アールグレイティー",
    "【1枚抜き専用】東洋科学様 アールグレイティー22(猫6柄)(W2)",
    "【1枚抜き専用】東洋科学様 アップルティー23(W2）",
    "【1枚抜き専用】東洋科学様 グリーンスムージー21",
    "【1枚抜き専用】東洋科学様 グリーンスムージーONEDAY21",
    "【1枚抜き専用】東洋科学様 ブラック無糖22秋",
    "【1枚抜き専用】東洋科学様 ブラック無糖22秋（HSOP）",
    "【1枚抜き専用】東洋科学様 マスカットティー",
    "【1枚抜き専用】東洋科学様 マスカットティー(HSOP)",
    "【1枚抜き専用】東洋科学様 マスカットティー23(W2）",
    "【1枚抜き専用】東洋科学様 マスカットティーRN23秋(W2）",
    "【1枚抜き専用】東洋科学様 レモンティー23(W2）",
    "【1枚抜き専用】東洋科学様 飲むパンケーキ",
    "【1枚抜き専用】東洋科学様 果肉ごろごろフルーツティー",
    "【ブッシュ専用】22秋Little Asia杏仁豆腐",
    "【ブッシュ専用】LWホットカフェラテ",
    "【ブッシュ専用】LWホットカフェラテ2023",
    "【ブッシュ専用】カラメルゼリーに恋したバタープリン",
    "【ブッシュ専用】サンプルラベル71-110-250(ｽﾄﾛｰﾏｰｸ入り)",
    "【ブッシュ専用】東洋科学様 TEBESﾄｰﾀﾙｴﾅｼﾞｰ22RN",
    "【ブッシュ専用】東洋科学様 TEBESﾋﾞｭｰﾃｨｰﾌﾟﾗｽ22RN",
    "【レーザー専用】東洋科学様 TEBESﾄｰﾀﾙｴﾅｼﾞｰ22RN",
    "【レーザー専用】東洋科学様 TEBESﾋﾞｭｰﾃｨｰﾌﾟﾗｽ22RN",
    "【レーザー専用】東洋科学様 TEBESミルキーピーチ24",
    "【レーザー専用】東洋科学様 TEBESミルキーピーチ24(表示改版)",
    "【レーザー専用】東洋科学様 TEBESミルキーホワイト24",
    "【レーザー専用】東洋科学様 TEBESミルキーホワイト24(表示改版)",
    "【レーザー専用】東洋科学様 UC果汁たっぷりﾊﾟｲﾝｼﾞｬｽﾐﾝﾃｨｰ(W2)",
    "【レーザー専用】東洋科学様 UC果汁ｵﾚﾝｼﾞﾐｯｸｽｼﾞｬｽﾐﾝﾃｨｰ(W2)",
    "【レーザー専用】東洋科学様 UCシトラスレモンティー23(W2)",
    "【レーザー専用】東洋科学様 UCはちみつ紅茶(W2）",
    "【レーザー専用】東洋科学様 UCブラック無糖22春",
    "【レーザー専用】東洋科学様 UCブラック無糖22春（HSOP)",
    "【レーザー専用】東洋科学様 UC果汁たっぷりピーチティー",
    "【レーザー専用】東洋科学様 UC果汁たっぷりフルーツティー",
    "【レーザー専用】東洋科学様 アールグレイティー23(W2)",
    "【レーザー専用】東洋科学様 アップルティー23(W2)",
    "【レーザー専用】東洋科学様 アップルティーRN23秋(W2)",
    "【レーザー専用】東洋科学様 いちごミックススムージー22(W2)",
    "【レーザー専用】東洋科学様 グリーンスムージー21",
    "【レーザー専用】東洋科学様 グリーンスムージー22",
    "【レーザー専用】東洋科学様 グリーンスムージー22（HSOP)",
    "【レーザー専用】東洋科学様 グリーンスムージー22(W2)",
    "【レーザー専用】東洋科学様 グリーンスムージー23(W2)",
    "【レーザー専用】東洋科学様 グリーンスムージーONEDAY21",
    "【レーザー専用】東洋科学様 グリーンスムージーONEDAY22",
    "【レーザー専用】東洋科学様 グリーンスムージーONEDAY22(W2)",
    "【レーザー専用】東洋科学様 グリーンスムージーONEDAY23(W2)",
    "【レーザー専用】東洋科学様 ｸﾞﾘｰﾝｽﾑｰｼﾞｰONEDAYRN23秋(W2)",
    "【レーザー専用】東洋科学様 グリーンスムージーRN23秋(W2)",
    "【レーザー専用】東洋科学様 ゴールデンダイヤモンドパイン",
    "【レーザー専用】東洋科学様 ｼﾄﾗｽﾚﾓﾝﾃｨｰ24(7柄)(W2)",
    "【レーザー専用】東洋科学様 ショコラバナナ",
    "【レーザー専用】東洋科学様 スリムスタイルスムージー",
    "【レーザー専用】東洋科学様 ｾﾞﾝｼｮｰ様いちごミルク(4柄)(HSOP)",
    "【レーザー専用】東洋科学様 ピーチウーロン23(8柄)(W2)",
    "【レーザー専用】東洋科学様 ピーチティー",
    "【レーザー専用】東洋科学様 ピーチルイボスティー(W2）",
    "【レーザー専用】東洋科学様 マスカットティーRN23秋(W2)",
    "【レーザー専用】東洋科学様 レモンティー23(W2)",
    "【レーザー専用】東洋科学様 飲む台湾ｽｲｰﾂ豆花(原材料表示変更)",
    "【レーザー専用】東洋科学様 飲む台湾ｽｲｰﾂ豆花23RN",
    "【レーザー専用】東洋科学様 鉄観音烏龍(W2）（6柄）",
    "【レーザー専用】東洋科学様UCﾀﾋﾟｵｶﾐﾙｸﾃｨｰ ﾋｸﾞﾁﾕｳｺ様ﾃﾞｻﾞｲﾝ(W2）",
    "71-205　CVS珈琲ゼリー",
    "24秋SWEET CAFE珈琲ゼリー<S>PP71φ-205M-IML",
    "24春SWEET CAFE珈琲ゼリー<S>PP71φ-205M-IML",
    "25春SWEET CAFEホワイトモカゼリー(PP71φｰ200M)",
    "25春SWEET CAFE珈琲ゼリー(PP71φｰ200M)",
    "RP71-205 23春SC珈琲ゼリー",
    "25春珈琲ゼリープリン",
    "RP71-205　24SC珈琲ゼリー(3柄キャンペーン品)",
    "RP71-205 24春SC珈琲ゼリー",
    "SC珈琲ゼリー",
    "RP71-205コーヒーゼリー22春",
    "RP71-240コーヒーゼリー22春", 
    "カスピ海ヨーグルト　プレーン（総厚52μ）23年冬",
    "カスピ海ヨーグルト　プレーン（総厚72μ）",
    "カスピ海ヨーグルト　プレーン21秋（総厚52μ）",
    "カスピ海ヨーグルト　プレーン21秋（総厚72μ）",
    "カスピ海ヨーグルト　脂肪０（20春）（総厚72μ）",
    "カスピ海ヨーグルト　脂肪０（総厚52μ）23年冬",
    "カスピ海ヨーグルト　脂肪ゼロ（21秋）（総厚52μ）",
    "カスピ海ヨーグルト　脂肪ゼロ（21秋）（総厚72μ）",
    "カスピ海ヨーグルト　脂肪０（総厚52μ）24年秋",
    "ジュレパルフェ",
    "ドトールカフェラテ71-250 22春",
    "ドトールカフェラテ71-250 22春(W2)",
    "ドトール砂糖不使用ラテ",
    "まるごとSOYカスピ海ヨーグルト24（52μ）",
    "まるごとSOYカスピ海ヨーグルト25（52μ）",
    "まるごとＳＯＹカスピ海ヨーグルト52μ",
    "まるごとＳＯＹカスピ海ヨーグルト52μ(白版①)",
    "まるごとＳＯＹカスピ海ヨーグルト72μ",
    "原反）お土産用　よつ葉バター　フタ　Ｔ３",
    "原反）お土産用 よつ葉バターカップ 20ハル",
    "原反）お土産用　発酵バター　フタ",
    "原反）お土産用発酵バターカップ 20ハル",
    "原反）はちみつバター ",
    "よつ葉バター",
    "発酵バター",
]


mask = (
    (df['7_品種コード'] == 232) &
    (df['6_得意先略称'].apply(lambda x: any(c in str(x) for c in customer_pattern)))&
    (df['11_品名'].apply(lambda x: any(pat in x for pat in product_pattern)))
)
df.loc[mask, '56_製品分類No'] = 6



#品種コード231で，得意先が伊藤園である製品を製品分類No7に設定する(伊藤園シュリンク)
mask = (df['7_品種コード'] == 231) & \
       (df['6_得意先略称']=='伊藤園')
df.loc[mask, '56_製品分類No'] = 7


##品種コードが231で，得意先がボンパック．品名がprod_listの製品について，製品分類Ｎｏを8に設定する(ボンパックシュリンク)
product_list = [
    "【試作】しぼりたて生200ｍｌ（HTO7）",
    "【試作】原反改訂品　新鮮しぼりたて生　450ml　ECO",
    "【試作】新鮮おさしみ生200ｍｌ(HTO7)",
    "【試作】味わいリッチ減塩200ｍｌ（BS551S）",
    "【試作】味わいリッチ減塩200ｍｌ（HNK）",
    "【試作】味わいリッチ減塩200ｍｌ（HT07）",
    "【試作】味わいリッチ減塩200ｍｌ（S7561）",
    "【試作】味わいリッチ減塩450ｍｌ　ミシン目変更品",
    "【試作2】味わいリッチ減塩200ｍｌ（HT07）",
    "2003  TV緑茶2L（まいばすけっと用)",
    "B】レモンスカッシュのもと350ml",
    "B】新鮮しぼりたて生　620ml",
    "TVMSジャスミン茶600ｍｌ",
    "TVMS麦茶600ｍｌ",
    "TVディズニー麦茶　500ｍｌ　トイストーリー",
    "TVディズニー緑茶　500ｍｌ　ミッキーフレンズ",
    "TV烏龍茶2L（まいばすけっと用)",
    "TV烏龍茶500ml(まいばすけっと用)",
    "TV緑茶500ｍｌ(まいばすけっと用)",
    "ｻﾝﾌﾟﾙ)しぼ生450ｼｭﾘﾝｸ25春夏RNﾃｽ",
    "しぼりたてうすくち生しょうゆ450ｍｌ",
    "しぼりたて生200ｍｌ（ロール）",
    "しぼりたて生醤油200ｍｌ　2021秋バージョン",
    "はちみつレモンスカッシュ　410ｍｌ",
    "ペコスカッシュ国産白桃　410ｍl　40μ",
    "ペコスカッシュ白桃　410ｍｌ",
    "メロンスカッシュ　410ｍｌ",
    "レモンコーラスカッシュ　410ｍｌ",
    "レモンスカッシュ　500ｍｌ",
    "塩分控えめ丸大豆450ｍｌ",
    "新鮮おさしみ生200ｍｌ",
    "新鮮おさしみ生200ｍｌ　お正月バージョン",
    "新鮮おさしみ生200ｍｌ　花火バージョン",
    "新鮮しぼりたて生　450ml",
    "新鮮しぼりたて生　450ml　1808",
    "新鮮しぼりたて生　450ml　2021秋バージョン",
    "新鮮しぼりたて生　450ml　ECO",
    "新鮮しぼりたて生　450ml　ECO　秋バージョン",
    "新鮮しぼりたて生　450ml　春バージョン",
    "新鮮しぼりたて生　620ml",
    "新鮮塩分ひかえめ丸大豆200ｍｌ",
    "新鮮丸大豆　生450ｍｌ",
    "新鮮超特選二段熟成　生しょうゆ450ｍｌ",
    "超特選極旨しょうゆ450ｍｌ",
    "都光　養老山地の天然水2L",
    "特選ﾘｯﾁ丸大豆　450ml",
    "特選ﾘｯﾁ丸大豆　620ml",
    "特選ﾘｯﾁ丸大豆（食べログ百名店スポット）450ml",
    "特選ﾘｯﾁ丸大豆（食べログ百名店スポット）620ml",
    "米麹純米本みりん450ｍｌ",
    "米麹本みりん450ｍｌ",
    "米麹本みりん450ｍｌ　秋バージョン",
    "米麹本みりん450ｍｌ　春バージョン",
    "米麹本みりん620ｍｌ",
    "宝石のレモンスカッシュルビー410ｍｌ 7410枚（40μ）",
    "本膳しょうゆ450ｍｌ",
    "味わいリッチ減塩　620ml",
    "味わいリッチ減塩200ｍｌ（ロール）",
    "味わいリッチ減塩450ｍｌ",
    "味わいリッチ減塩450ｍｌ　秋バージョン",
    "味わいリッチ減塩450ｍｌ春バージョン",
    "輸出新鮮おさしみDF200ｍｌ",
    "輸出新鮮しぼりたて生SA200ｍｌ",
]
mask = (
    (df['7_品種コード'] == 231) &
    (df['6_得意先略称'] == 'ボンパック') &
    df['11_品名'].apply(lambda x: any(pat in x for pat in product_list))
)
df.loc[mask, '56_製品分類No'] =8





# 未定義の製品分類Noについて、条件分けして番号を付けていく
# 条件1: 品種コードが230で、製造分類Noが振られていない製品の番号を９とする
mask = (
    (df['7_品種コード'] == 230) &
    df['56_製品分類No'].isna()
)
df.loc[mask, '56_製品分類No'] = 9

# 条件2: 品種コードが236以外の場合
mask = (
    (df['7_品種コード'] == 236) &
    df['56_製品分類No'].isna()
)
df.loc[mask, '56_製品分類No'] = 10

# 条件2: 品種コードが236以外の場合
mask = (
    (df['7_品種コード'] == 232) &
    df['56_製品分類No'].isna()
)
df.loc[mask, '56_製品分類No'] = 11


# 条件2: 品種コードが236以外の場合
mask = (
    (df['7_品種コード'] == 231) &
    df['56_製品分類No'].isna()
)
df.loc[mask, '56_製品分類No'] = 12


# 条件2: 品種コードが236以外の場合
mask = (
    (df['7_品種コード'] == 233) &
    df['56_製品分類No'].isna()
)
df.loc[mask, '56_製品分類No'] = 13

# 未定義の製品分類Noを9に設定
mask = df['56_製品分類No'].isna()
df.loc[mask, '56_製品分類No'] = 14



# 製品分類Noと製品分類名の対応付け
product_category_map = {
    1: '麦茶54p',
    2: 'ZIP',
    3: '食品トレー',
    4: '伊藤園製袋',
    5: '伊藤園ロール',
    6: 'ＩＭＬ',
    7: '伊藤園シュリンク',
    8: 'ボンパックシュリンク',
    9: 'その他軟材(製袋)',
    10:'その他軟材(ロール)',
    11:'その他インモールドラベル',
    12:'その他シュリンク',
    13:'その他ロールラベル',
    14:'その他',
}

# 製品分類名のカラムを追加
df['57_製品分類名'] = df['56_製品分類No'].map(product_category_map)






# Streamlitのタイトルと説明
st.title("過去実績の全体分析アプリ")
st.write("機械名を入力して検索結果を表示します。")

# 印刷と記入の選択肢を追加
search_type = st.radio(
    "製造工程を選択してください",
    ["印刷", "ドライラミ", "押し出しラミ"],
    key="search_type_radio"
)

# 確定ボタンを追加
confirm_button1 = st.button("選択を確定", key="confirm_button_1")

# 検索タイプに基づいて行を抽出し、行数を表示
if confirm_button1:
    if search_type == "印刷":
        # UW-1,UW-3,UW-5を含む機械名を抽出
        filtered_df = df[df['3_機械名'].str.contains('ＵＷ－１|ＵＷ－３|ＵＷ－５|ＵＷ－７', na=False)]

    elif search_type == "ドライラミ":
        # UD-2, UD-3を含む機械名を抽出
        filtered_df = df[df['3_機械名'].str.contains('ＵＤ－２|ＵＤ－３', na=False)]

    elif search_type == "押し出しラミ":
        # UE-1を含む機械名を抽出
        filtered_df = df[df['3_機械名'].str.contains('UE-1', na=False)]

    # 状態を保持
    st.session_state.filtered_df = filtered_df

# filtered_dfが存在することを確認
if 'filtered_df' in st.session_state:
    filtered_df = st.session_state.filtered_df
    st.write(f"### 抽出されたデータ数: {len(filtered_df)}行")
    if not filtered_df.empty:
        st.write("データが正常に抽出されました。")

        # 日付範囲の選択を追加
        st.write("### 期間を選択してください")

        # データの日付範囲を取得
        min_date = filtered_df['4_製造日'].min()
        max_date = filtered_df['4_製造日'].max()

        col1, col2 = st.columns(2)

        with col1:
            start_date = st.date_input(
                "開始日",
                value=min_date,
                min_value=min_date,
                max_value=max_date
            )

        with col2:
            end_date = st.date_input(
                "終了日",
                value=max_date,
                min_value=min_date,
                max_value=max_date
            )

        # 確定ボタンを追加
        confirm_date = st.button("日付範囲を確定", key="confirm_date")

        if confirm_date:
            # 日付範囲でフィルタリング
            filtered_df['4_製造日'] = pd.to_datetime(filtered_df['4_製造日'])
            filtered_df = filtered_df[
                (filtered_df['4_製造日'] >= pd.to_datetime(start_date)) & 
                (filtered_df['4_製造日'] <= pd.to_datetime(end_date))
            ]
            st.session_state.filtered_df = filtered_df
            st.write("日付範囲でフィルタリングしました。")
    else:
        st.write("条件に一致するデータが見つかりませんでした。")
else:
    st.write("先に製造工程を選択してください。")

# 期間で絞った後のfiltered_dfが定義されていることを確認
if 'filtered_df' in st.session_state and not st.session_state.filtered_df.empty:
    filtered_df = st.session_state.filtered_df

    # 日付範囲でフィルタリングされたデータの確認
    st.write("### 期間でフィルタリングされたデータ")
    st.write(f"抽出されたデータ数: {len(filtered_df)}行")

    # 日付範囲の表示
    min_date = filtered_df['4_製造日'].min()
    max_date = filtered_df['4_製造日'].max()
    st.write(f"期間: {min_date} から {max_date}")

    # データのプレビュー
    st.write("### データプレビュー")
    st.dataframe(filtered_df.head())
else:
    st.write("条件に一致するデータが見つかりませんでした。")


# 製造工程が選択され、filtered_dfが定義されていることを確認
if 'filtered_df' in st.session_state and not filtered_df.empty:
    # 主要品目の選択肢を追加
    search_type_2 = st.multiselect(
        "主要品目を選択してください（複数選択可）",
        ["すべて", "麦茶54p", "ZIP", "食品トレー", "伊藤園製袋", "伊藤園ロール", "ＩＭＬ", '伊藤園シュリンク', 'ボンパックシュリンク', 'その他軟材(製袋)', 'その他軟材(ロール)', 'その他インモールドラベル', 'その他シュリンク', 'その他ロールラベル', 'その他'],
        key="search_type_multiselect_2"
    )


    # 確定ボタンを追加
    confirm_button2 = st.button("選択を確定", key="confirm_button_2")

    # 製品分類のマッピングを定義
    product_mapping = {
        "麦茶54p": 1,
        "ZIP": 2,
        "食品トレー": 3,
        "伊藤園製袋": 4,
        "伊藤園ロール": 5,
        "ＩＭＬ": 6,
        "伊藤園シュリンク": 7,
        "ボンパックシュリンク": 8,
        "その他軟材(製袋)": 9,
        "その他軟材(ロール)": 10,
        "その他インモールドラベル": 11,
        "その他シュリンク": 12,
        "その他ロールラベル": 13,
        "その他": 14
    }

    if confirm_button2:
        filtered_dfs = []
        
        # 'すべて' が選択されている場合
        if "すべて" in search_type_2:
            filtered_dfs.append(filtered_df)  # すべてのデータを追加
            
            # その他の選択肢も処理
            other_selections = [item for item in search_type_2 if item != "すべて"]
            if other_selections:
                selected_numbers = [product_mapping.get(item) for item in other_selections]
                for number in selected_numbers:
                    filtered_dfs.append(filtered_df[filtered_df['56_製品分類No'] == number])
        else:
            # 'すべて' 以外の選択肢のみの場合
            selected_numbers = [product_mapping.get(item) for item in search_type_2]
            for number in selected_numbers:
                filtered_dfs.append(filtered_df[filtered_df['56_製品分類No'] == number])

# filtered_dfsが定義されているか確認
if 'filtered_dfs' not in locals() or not filtered_dfs:
    st.error("❌ 主要品目が選択されていません。先に主要品目を選択してください。")
    st.stop()



# 更新対象列の定義
columns_to_update = [
    '18_投入数', '19_製造数', '20_ロス',
    '21_所要時間準備', '22_所要時間調色', '23_所要時間計画', '24_所要時間突発', '25_所要時間立合',
    '26_所要時間工程', '27_所要時間後始末', '28_ロス時間合計', '29_所要時間基礎', '30_所要時間合計',
    '33_調色回数' , '34_手配数', '35_責任完了数' ,'37_巻M数1', '38_巻数1', '39_巻M数2', '40_巻数2',
    '41_通し（総）ｍ数','42_生産金額+荷作り代', '50_生産金額', '51_荷作り代', '52_ロスM数', '53_原反巾'
]

columns_to_standardize = [
    '1_工場区分名', '2_機械工程分類区分名', '3_機械名', '4_製造日', '5_受注区分名', '6_得意先略称', '7_品種コード',
    '8_原反コード', '9_原反名称', '10_受注№', '11_品名', '12_版区分', '13_工程コード', '14_工程名',
    '15_作業者', '16_操業区分', '17_完了区分', '32_色数', '44_ラミ工程', '46_品種分類コード', '47_業種コード',
    '48_印刷ロス率(受注)', '49_印刷ロス率(実績)','54_原反受入単価', '55_工程単価','43_印刷速度',  '45_ラミ速度'
]

# filtered_dfsの各データフレームに対して処理を実行
processed_dfs = []
for df_categoryi in filtered_dfs:
    df_categoryi = df_categoryi.copy()
    df_categoryi['4_製造日'] = pd.to_datetime(df_categoryi['4_製造日'], errors='coerce').dt.date
    df_categoryi = df_categoryi.reset_index(drop=True)
    df_categoryi.index = df_categoryi.index + 1

    # 準備と後始末の数値データを稼働した行に結合
    # パターンA: 稼働を準備と後始末ではさんでいる場合
    df_categoryi['prev_worker'] = df_categoryi['15_作業者'].shift(1)
    df_categoryi['next_worker'] = df_categoryi['15_作業者'].shift(-1)
    mask = (df_categoryi['prev_worker'].str.contains('準備', na=False)) & \
           (df_categoryi['next_worker'].str.contains('後始末', na=False))
    for column in columns_to_update:
        df_categoryi.loc[mask, column] += df_categoryi[column].shift(1)[mask] + df_categoryi[column].shift(-1)[mask]
    df_categoryi.drop(['prev_worker', 'next_worker'], axis=1, inplace=True)

    # パターンB: 準備だけ先にやられている場合
    df_categoryi['prev_worker'] = df_categoryi['15_作業者'].shift(1)
    df_categoryi['next_worker'] = df_categoryi['15_作業者'].shift(-1)
    mask = (df_categoryi['prev_worker'].str.contains('準備', na=False)) & \
           (~df_categoryi['next_worker'].str.contains('後始末', na=False))
    for column in columns_to_update:
        df_categoryi.loc[mask, column] += df_categoryi[column].shift(1)[mask]
    df_categoryi.drop(['prev_worker', 'next_worker'], axis=1, inplace=True)

    # パターンC: 後始末だけ別で後に実施した場合
    df_categoryi['prev_worker'] = df_categoryi['15_作業者'].shift(1)
    df_categoryi['next_worker'] = df_categoryi['15_作業者'].shift(-1)
    mask = (~df_categoryi['prev_worker'].str.contains('準備', na=False)) & \
           (df_categoryi['next_worker'].str.contains('後始末', na=False))
    for column in columns_to_update:
        df_categoryi.loc[mask, column] += df_categoryi[column].shift(-1)[mask]
    df_categoryi.drop(['prev_worker', 'next_worker'], axis=1, inplace=True)

    # 準備と後始末の行を削除
    df_categoryi = df_categoryi[~df_categoryi['15_作業者'].str.contains('準備|後始末|中止', na=False)]
    df_categoryi = df_categoryi.reset_index(drop=True)
    df_categoryi.index = df_categoryi.index + 1

    # 稼働番号処理
    df_categoryi['操業区分'] = df_categoryi['16_操業区分'].apply(lambda x: '夜' if '夜' in str(x) else '昼')
    df_categoryi['稼働番号'] = 1
    for j in range(2, len(df_categoryi) + 1):
        prev_date = df_categoryi.loc[j-1, '4_製造日']
        prev_product = df_categoryi.loc[j-1, '11_品名']
        prev_shift = df_categoryi.loc[j-1, '操業区分']
        prev_a = df_categoryi.loc[j-1, '稼働番号']
        curr_date = df_categoryi.loc[j, '4_製造日']
        curr_product = df_categoryi.loc[j, '11_品名']
        curr_shift = df_categoryi.loc[j, '操業区分']
        if curr_date == prev_date and curr_product == prev_product and curr_shift != prev_shift:
            df_categoryi.loc[j, '稼働番号'] = prev_a
        elif pd.to_datetime(curr_date) == pd.to_datetime(prev_date) + pd.Timedelta(days=1) and \
             curr_product == prev_product and prev_shift == '夜' and curr_shift == '昼':
            df_categoryi.loc[j, '稼働番号'] = prev_a
        else:
            df_categoryi.loc[j, '稼働番号'] = prev_a + 1

    # 各稼働番号の1行目データ・合計を取得
    first_values = df_categoryi.groupby('稼働番号')[columns_to_standardize].first()
    for col in columns_to_standardize:
        df_categoryi[col] = df_categoryi['稼働番号'].map(first_values[col])

    sum_values = df_categoryi.groupby('稼働番号')[columns_to_update].sum()
    for col in columns_to_update:
        df_categoryi[col] = df_categoryi['稼働番号'].map(sum_values[col])

    df_categoryi = df_categoryi.groupby('稼働番号').first().reset_index()

    # ロス・率などの計算
    df_categoryi = df_categoryi.rename(columns={'20_ロス': '20_製造ロス'})
    df_categoryi['品質ロス'] = np.where(df_categoryi['20_製造ロス'] >= 150, df_categoryi['20_製造ロス'] - 150, 0)
    df_categoryi['製造ロス率'] = ((df_categoryi['20_製造ロス'] / df_categoryi['18_投入数']) * 100).round(1)
    df_categoryi['品質ロス率'] = ((df_categoryi['品質ロス'] / df_categoryi['18_投入数']) * 100).round(1)
    df_categoryi['リード紙ロス率'] = ((df_categoryi['36_リード紙Ｍ数'] / df_categoryi['18_投入数']) * 100).round(1)
    df_categoryi['実績ロス合計'] = (df_categoryi['製造ロス率'] + df_categoryi['リード紙ロス率']).round(1)
    df_categoryi['工程稼働率'] = ((df_categoryi['26_所要時間工程'] / df_categoryi['30_所要時間合計']) * 100).round(1)

    processed_dfs.append(df_categoryi)

# 処理済みのデータフレームを結合
final_df = pd.concat(processed_dfs, ignore_index=True)

columns = [
    'ロットサイズ', 'ロットサイズ_最小', 'ロットサイズ_最大', '稼働数',  '製造数量', '平均数量',
    '数量ばらつき', '所要時間合計','所要時間基礎','所要時間工程','工程稼働率（%）', '工程稼働率ばらつき', '調色時間', '調色時間ばらつき',
    '平均ロスm数', '平均品質ロスm数', 'リードしM数', '製造ロス率（%）', '品質ロス率（%）', 'リード紙ロス率（%）', '品質ロスばらつき'
]

lot_size_data = [
    ['全体', 0, 99999999],
    ['60000-999999', 60000, 99999999],
    ['32000-60000', 32000, 60000],
    ['8000-32000', 8000, 32000],
    ['4000-8000', 4000, 8000],
    ['4000未満', 0, 4000]
]

all_results = []
results_dfs = {}  # 空の辞書を用意
roop=0
# 各リストごとの処理
for df in processed_dfs:
    results = []
    roop+=1
    
    # ロットサイズごとの処理
    for data in lot_size_data:
        min_val, max_val = data[1], data[2]
        filtered = df[(df['19_製造数'] > min_val) & (df['19_製造数'] <= max_val)]
        
        # メトリクス計算
        kado = len(filtered)
        seizo_total = filtered['19_製造数'].sum()
        avg_qty = int(seizo_total / kado) if kado else np.nan
        
        # 3桁ごとにコンマを打つ
        kado = f"{kado:,}"
        seizo_total = f"{seizo_total:,}"
        avg_qty = f"{avg_qty:,}" if not np.isnan(avg_qty) else np.nan

        mean = filtered['19_製造数'].mean()
        std = filtered['19_製造数'].std()
        cv_qty = std / mean if mean else np.nan
        cv_qty = round(cv_qty, 2)

        total_26 = filtered['26_所要時間工程'].sum()
        total_29=filtered['29_所要時間基礎'].sum()
        total_30 = filtered['30_所要時間合計'].sum()
        kadoritsu = (total_26 / total_30 * 100).round(2) if total_30 else np.nan

    
        rates = filtered['26_所要時間工程'] / filtered['30_所要時間合計']
        rates = (rates * 100).round(2)
        mean_r = rates.mean()
        std_r = rates.std()
        kadoritsu_cv = std_r / mean_r if mean_r else np.nan
        kadoritsu_cv = round(kadoritsu_cv, 2)


        choshoku_mean = filtered['22_所要時間調色'].mean()
        choshoku_mean = round(choshoku_mean, 1)
        choshoku_std = filtered['22_所要時間調色'].std()
        choshoku_cv = choshoku_std / choshoku_mean if choshoku_mean else np.nan
        choshoku_cv=round(choshoku_cv,2)

        loss_avg = round(filtered['20_製造ロス'].mean(), 1)
        q_loss_avg = round(filtered['品質ロス'].mean(), 1)
        lead_avg = round(filtered['36_リード紙Ｍ数'].mean(), 1)

        

        with pd.option_context('mode.use_inf_as_na', True):
            loss_rate = ((filtered['20_製造ロス'] / filtered['18_投入数']) * 100).mean()
            loss_rate = round(loss_rate, 2) if pd.notnull(loss_rate) else np.nan

            q_loss_rate = ((filtered['品質ロス'] / filtered['18_投入数']) * 100).mean()
            q_loss_rate = round(q_loss_rate, 2) if pd.notnull(q_loss_rate) else np.nan

            lead_rate = ((filtered['36_リード紙Ｍ数'] / filtered['18_投入数']) * 100).mean()
            lead_rate = round(lead_rate, 2) if pd.notnull(lead_rate) else np.nan

        q_loss_std = filtered['品質ロス'].std()
        q_loss_cv = round(q_loss_std / q_loss_avg if q_loss_avg else np.nan,2)

        metrics = [int(kado.replace(',', '')), int(seizo_total.replace(',', '')), int(avg_qty.replace(',', '')) if not pd.isna(avg_qty) else np.nan,
              cv_qty, total_30, total_29, total_26,
                  kadoritsu, kadoritsu_cv, choshoku_mean, choshoku_cv,
                  loss_avg, q_loss_avg, lead_avg,
                  loss_rate, q_loss_rate, lead_rate, q_loss_cv]
        
        results.append([data[0], min_val, max_val] + metrics)

    # 試作品の処理
    filtered = df[df['5_受注区分名'].isin(['再生産', '試作(無償)', '試作(有償)'])]
    
    # メトリクス計算
    kado = len(filtered)
    seizo_total = filtered['19_製造数'].sum()
    avg_qty = int(seizo_total / kado) if kado else np.nan
        
        # 3桁ごとにコンマを打つ
    kado = f"{kado:,}"
    seizo_total = f"{seizo_total:,}"
    avg_qty = f"{avg_qty:,}" if not np.isnan(avg_qty) else np.nan

    mean = filtered['19_製造数'].mean()
    std = filtered['19_製造数'].std()
    cv_qty = std / mean if mean else np.nan
    cv_qty = round(cv_qty, 2)
    total_26 = filtered['26_所要時間工程'].sum()
    total_29=filtered['29_所要時間基礎'].sum()
    total_30 = filtered['30_所要時間合計'].sum()
    kadoritsu = (total_26 / total_30 * 100).round(2) if total_30 else np.nan

    rates = filtered['26_所要時間工程'] / filtered['30_所要時間合計']
    rates = (rates * 100).round(2)
    mean_r = rates.mean()
    std_r = rates.std()
    kadoritsu_cv = std_r / mean_r if mean_r else np.nan
    kadoritsu_cv = round(kadoritsu_cv, 2)

    choshoku_mean = filtered['22_所要時間調色'].mean()
    choshoku_mean = round(choshoku_mean, 1)
    choshoku_std = filtered['22_所要時間調色'].std()
    choshoku_cv = choshoku_std / choshoku_mean if choshoku_mean else np.nan
    choshoku_cv=round(choshoku_cv,2)
    loss_avg = round(filtered['20_製造ロス'].mean(), 1)
    q_loss_avg = round(filtered['品質ロス'].mean(), 1)
    lead_avg = round(filtered['36_リード紙Ｍ数'].mean(), 1)

        

    with pd.option_context('mode.use_inf_as_na', True):
        loss_rate = ((filtered['20_製造ロス'] / filtered['18_投入数']) * 100).mean()
        loss_rate = round(loss_rate, 2) if pd.notnull(loss_rate) else np.nan        
        q_loss_rate = ((filtered['品質ロス'] / filtered['18_投入数']) * 100).mean()
        q_loss_rate = round(q_loss_rate, 2) if pd.notnull(q_loss_rate) else np.nan
        lead_rate = ((filtered['36_リード紙Ｍ数'] / filtered['18_投入数']) * 100).mean()
        lead_rate = round(lead_rate, 2) if pd.notnull(lead_rate) else np.nan

        q_loss_std = filtered['品質ロス'].std()
        q_loss_cv = round(q_loss_std / q_loss_avg if q_loss_avg else np.nan,2)

    metrics = [int(kado.replace(',', '')), int(seizo_total.replace(',', '')), int(avg_qty.replace(',', '')) if not pd.isna(avg_qty) else np.nan,
              cv_qty, total_30, total_29, total_26,
              kadoritsu, kadoritsu_cv, choshoku_mean, choshoku_cv,
              loss_avg, q_loss_avg, lead_avg,
              loss_rate, q_loss_rate, lead_rate, q_loss_cv]
    
    results.append(['試作', np.nan, np.nan] + metrics)
    all_results.extend(results)

    # 結果を表形式で表示
    st.write("### 分析結果の詳細")
    results_df = pd.DataFrame(results, columns=['ロットサイズ', '最小', '最大'] + columns[3:])
    # 製品分類名を1列目に追加（すべての行に同じ製品分類名を設定）
    results_df.insert(0, '製品分類', [search_type_2[roop-1]] * len(results_df))
    st.dataframe(results_df)
    
     # i-1に注意
    results_dfs[roop] = results_df # 数字キーで格納（例：1, 2, 3...）
 

    
    


df2 = pd.DataFrame(all_results, columns=columns)



# 複数選択の場合の処理
if len(search_type_2) > 1:
    # 全体の行を抽出
    overall_rows = df2[df2['ロットサイズ'] == '全体'].copy()
    
    # 製品分類名を追加
    overall_rows.insert(0, '製品分類', search_type_2)
    
    # 結果を表示
    st.write("### 複数選択時の分析結果（全体のみ）")
    st.dataframe(overall_rows)
# df2とoverall_rowsの内容を確認
st.write("### データフレームの内容確認")



st.write("### 分析結果の保存")

# 入力フィールド（シート名）
sheet_name = st.text_input("保存する詳細分析結果のシート名", value="分析結果")
overall_sheet_name = st.text_input("保存する全体分析結果のシート名", value="全体分析結果")

# ファイルがアップロードされていない場合のエラーメッセージ
if uploaded_file is None:
    st.error("❌ エクセルファイルをアップロードしてください")
else:
    # 一時保存パスを定義
    temp_path = f"temp_{original_filename}"
    
    # ワークブックを読み込み
    book = openpyxl.load_workbook(temp_path)

    # 既存シートの削除
    for target_sheet in [sheet_name, overall_sheet_name if len(search_type_2) > 1 else None]:
        if target_sheet and target_sheet in book.sheetnames:
            del book[target_sheet]

    # 削除後に保存
    book.save(temp_path)

    # 各データフレームを結合するための準備
    combined_df = pd.DataFrame()
    
    # 各データフレームを順番に結合
    for i in range(1, roop + 1):
        if i > 1:
            # 2つ目以降のデータフレームの場合、空行を追加
            empty_row = pd.DataFrame([[''] * len(combined_df.columns)], columns=combined_df.columns)
            combined_df = pd.concat([combined_df, empty_row], ignore_index=True)
        
        # 現在のデータフレームを追加
        combined_df = pd.concat([combined_df, results_dfs[i]], ignore_index=True)
    
    # 結合したデータフレームを保存
    results_dfs['combined'] = combined_df
    
    
    # ExcelWriterで新しいシートを追加
    with pd.ExcelWriter(temp_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 全体分析結果（複数選択時のみ）
        if len(search_type_2) > 1:
            overall_rows.to_excel(writer, sheet_name=overall_sheet_name, index=False)

    # 成功メッセージと情報表示
    st.success("✅ 分析結果を保存しました")
    st.write(f"保存先ファイル: `{temp_path}`")
    st.write(f"📄 詳細分析結果シート: `{sheet_name}`")
    if len(search_type_2) > 1:
        st.write(f"📄 全体分析結果シート: `{overall_sheet_name}`")

    # セッション状態の更新
    st.session_state.saved = True
    st.session_state.continue_analysis = True

    # ダウンロードボタンの表示
    with open(temp_path, 'rb') as f:
        st.download_button(
            label="📥 保存したファイルをダウンロード",
            data=f,
            file_name=original_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

